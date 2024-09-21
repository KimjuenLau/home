import os
from openpyxl import load_workbook

def get_filenames_without_extension(folder_path):
    filenames_without_extension = []
    for filename in os.listdir(folder_path):
        if os.path.isfile(os.path.join(folder_path, filename)):
            filenames_without_extension.append(str(os.path.splitext(filename)[0]))
    return filenames_without_extension

def get_non_empty_cells_from_A_column(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    result = []
    
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
        cell_value = str(row[0])
        if cell_value is None:
            break
        result.append(cell_value)
    
    return result

def get_common_element_indices(list1, list2):
    common_indices = []
    
    for index, element in enumerate(list1):
        if element in list2:
            common_indices.append(index)
    
    return common_indices

def rename_file(src, dst):
    try:
        os.rename(src, dst)
        print(f"文件重命名成功：'{src}' -> '{dst}'")
    except FileNotFoundError:
        print(f"未找到文件：{src}")
    except PermissionError:
        print("权限不足，无法重命名文件。")
    except Exception as e:
        print(f"发生错误：{e}")



file_path_folder = 'E:\Echoes of Home'
file_path_excel = 'E:\Echoes of Home\新建 XLSX 工作表.xlsx'
a = get_filenames_without_extension(file_path_folder)
b = get_non_empty_cells_from_A_column(file_path_excel)
for i in get_common_element_indices(a, b):
    rename_file(os.path.join(file_path_folder, a[i]+'.docx'), os.path.join(file_path_folder, "Used-"+a[i]+'.docx'))
print(a,b)
